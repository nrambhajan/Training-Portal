from datetime import datetime
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
import models, schemas
from auth import require_admin, get_current_user

router = APIRouter(prefix="/dashboard", tags=["dashboard"])

# ── Helper: gather trainee report data ──────────────────────────────────────

def _trainee_report_data(trainee_id: int, db: Session):
    trainee = db.query(models.User).filter(
        models.User.id == trainee_id, models.User.role == "trainee"
    ).first()
    if not trainee:
        raise HTTPException(status_code=404, detail="Trainee not found")

    modules = db.query(models.Module).order_by(models.Module.id).all()
    rows = []
    total_earned = 0.0
    total_max = 0.0
    total_correct = 0
    total_questions = 0

    for m in modules:
        q_count = len(m.questions)
        m_earned = 0.0
        m_max = sum(q.points for q in m.questions)
        m_correct = 0

        for q in m.questions:
            att = (
                db.query(models.Attempt)
                .filter(models.Attempt.user_id == trainee_id, models.Attempt.question_id == q.id)
                .order_by(models.Attempt.attempted_at.desc())
                .first()
            )
            if att and att.is_correct:
                m_correct += 1
                m_earned += att.score or 0

        pct = round((m_earned / m_max * 100) if m_max else 0, 1)
        rows.append({
            "day": rows.__len__() + 1,
            "title": m.title,
            "questions": q_count,
            "correct": m_correct,
            "earned": round(m_earned, 1),
            "max": round(m_max, 1),
            "percent": pct,
            "status": "Pass" if pct >= 60 else "Fail",
        })
        total_earned += m_earned
        total_max += m_max
        total_correct += m_correct
        total_questions += q_count

    return {
        "trainee": trainee,
        "rows": rows,
        "total_earned": round(total_earned, 1),
        "total_max": round(total_max, 1),
        "total_correct": total_correct,
        "total_questions": total_questions,
        "total_percent": round((total_earned / total_max * 100) if total_max else 0, 1),
    }


@router.get("/overview")
def overview(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    trainees = db.query(models.User).filter(models.User.role == "trainee").all()
    modules = db.query(models.Module).all()
    all_questions = db.query(models.Question).all()
    total_questions = len(all_questions)

    result = []
    for trainee in trainees:
        earned = 0.0
        max_pts = sum(q.points for q in all_questions)
        correct = 0
        attempted = 0

        for q in all_questions:
            att = (
                db.query(models.Attempt)
                .filter(
                    models.Attempt.user_id == trainee.id,
                    models.Attempt.question_id == q.id,
                )
                .order_by(models.Attempt.attempted_at.desc())
                .first()
            )
            if att:
                attempted += 1
                if att.is_correct:
                    correct += 1
                    earned += att.score or 0

        result.append({
            "trainee_id": trainee.id,
            "trainee_name": trainee.full_name or trainee.username,
            "username": trainee.username,
            "server_ip": trainee.server_ip,
            "total_questions": total_questions,
            "attempted": attempted,
            "correct": correct,
            "score": round(earned, 2),
            "max_score": round(max_pts, 2),
            "percent": round((earned / max_pts * 100) if max_pts else 0, 1),
        })

    return {
        "trainees": result,
        "total_modules": len(modules),
        "total_questions": total_questions,
        "total_trainees": len(trainees),
    }


@router.get("/trainee/{trainee_id}")
def trainee_detail(
    trainee_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    trainee = db.query(models.User).filter(
        models.User.id == trainee_id, models.User.role == "trainee"
    ).first()
    if not trainee:
        return {"error": "Trainee not found"}

    modules = db.query(models.Module).all()
    module_data = []

    for m in modules:
        questions_data = []
        for q in m.questions:
            att = (
                db.query(models.Attempt)
                .filter(
                    models.Attempt.user_id == trainee_id,
                    models.Attempt.question_id == q.id,
                )
                .order_by(models.Attempt.attempted_at.desc())
                .first()
            )
            questions_data.append({
                "question_id": q.id,
                "type": q.type,
                "text": q.text,
                "points": q.points,
                "correct_answer": q.correct_answer,
                "attempt": {
                    "id": att.id,
                    "submitted_answer": att.submitted_answer,
                    "server_output": att.server_output,
                    "is_correct": att.is_correct,
                    "score": att.score,
                    "attempted_at": att.attempted_at.isoformat(),
                    "graded_at": att.graded_at.isoformat() if att.graded_at else None,
                    "admin_notes": att.admin_notes,
                } if att else None,
            })
        module_data.append({
            "module_id": m.id,
            "module_title": m.title,
            "questions": questions_data,
        })

    return {
        "trainee": {
            "id": trainee.id,
            "username": trainee.username,
            "full_name": trainee.full_name,
            "server_ip": trainee.server_ip,
        },
        "modules": module_data,
    }


@router.put("/grade/{attempt_id}")
def grade_attempt(
    attempt_id: int,
    payload: schemas.GradeAttempt,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    attempt = db.query(models.Attempt).filter(models.Attempt.id == attempt_id).first()
    if not attempt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attempt not found")

    question = db.query(models.Question).filter(models.Question.id == attempt.question_id).first()

    attempt.is_correct = payload.is_correct
    attempt.score = question.points if payload.is_correct else 0.0
    attempt.graded_at = datetime.utcnow()
    if payload.admin_notes is not None:
        attempt.admin_notes = payload.admin_notes
    db.commit()
    db.refresh(attempt)

    return {
        "id": attempt.id,
        "is_correct": attempt.is_correct,
        "score": attempt.score,
        "graded_at": attempt.graded_at.isoformat(),
        "admin_notes": attempt.admin_notes,
    }


# ── PDF Report ──────────────────────────────────────────────────────────────

@router.get("/report/{trainee_id}/pdf")
def report_pdf(
    trainee_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    from fpdf import FPDF

    def _clean(text):
        """Replace unicode chars that Helvetica can't handle."""
        return (text or "").replace("\u2014", "-").replace("\u2013", "-").replace("\u2018", "'").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"').replace("\u2026", "...")

    d = _trainee_report_data(trainee_id, db)
    t = d["trainee"]
    now = datetime.utcnow().strftime("%d %b %Y, %H:%M UTC")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # --- Header with logo ---
    logo_path = Path(__file__).resolve().parent.parent.parent / "frontend" / "public" / "logo.svg"
    # fpdf2 doesn't support SVG natively; check for PNG fallback
    logo_png = Path(__file__).resolve().parent.parent / "logo.png"
    if logo_png.exists():
        pdf.image(str(logo_png), x=10, y=10, w=40)
        pdf.set_xy(55, 10)
    else:
        pdf.set_xy(10, 10)

    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, "Training Completion Report", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, "SupportSages - Linux Training Portal", ln=True)
    pdf.cell(0, 6, f"Generated: {now}", ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(8)

    # --- Trainee Info ---
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Trainee Information", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(50, 6, "Name:", 0)
    pdf.cell(0, 6, _clean(t.full_name or t.username), ln=True)
    pdf.cell(50, 6, "Username:", 0)
    pdf.cell(0, 6, _clean(t.username), ln=True)
    if t.server_ip:
        pdf.cell(50, 6, "Training Server:", 0)
        pdf.cell(0, 6, t.server_ip, ln=True)
    pdf.ln(6)

    # --- Overall Score ---
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Overall Performance", ln=True)
    pdf.set_font("Helvetica", "", 11)
    overall_status = "PASS" if d["total_percent"] >= 60 else "FAIL"
    pdf.cell(50, 7, "Total Score:", 0)
    pdf.cell(0, 7, f"{d['total_earned']} / {d['total_max']}  ({d['total_percent']}%)", ln=True)
    pdf.cell(50, 7, "Questions Correct:", 0)
    pdf.cell(0, 7, f"{d['total_correct']} / {d['total_questions']}", ln=True)
    pdf.cell(50, 7, "Result:", 0)
    if overall_status == "PASS":
        pdf.set_text_color(0, 128, 0)
    else:
        pdf.set_text_color(200, 0, 0)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, overall_status, ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(6)

    # --- Module Breakdown Table ---
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "Module-wise Breakdown", ln=True)
    pdf.ln(2)

    # Table header
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(40, 40, 40)
    pdf.set_text_color(255, 255, 255)
    col_w = [12, 68, 20, 20, 25, 18, 18]
    headers = ["Day", "Module", "Qs", "Correct", "Score", "%", "Status"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 8, h, 1, 0, "C", fill=True)
    pdf.ln()

    # Table rows
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)
    for i, r in enumerate(d["rows"]):
        if i % 2 == 0:
            pdf.set_fill_color(245, 245, 245)
        else:
            pdf.set_fill_color(255, 255, 255)

        pdf.cell(col_w[0], 7, str(r["day"]), 1, 0, "C", fill=True)
        # Truncate long titles
        title = _clean(r["title"])
        if len(title) > 35:
            title = title[:32] + "..."
        pdf.cell(col_w[1], 7, title, 1, 0, "L", fill=True)
        pdf.cell(col_w[2], 7, str(r["questions"]), 1, 0, "C", fill=True)
        pdf.cell(col_w[3], 7, str(r["correct"]), 1, 0, "C", fill=True)
        pdf.cell(col_w[4], 7, f"{r['earned']}/{r['max']}", 1, 0, "C", fill=True)
        pdf.cell(col_w[5], 7, f"{r['percent']}%", 1, 0, "C", fill=True)

        # Color the status
        if r["status"] == "Pass":
            pdf.set_text_color(0, 128, 0)
        else:
            pdf.set_text_color(200, 0, 0)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(col_w[6], 7, r["status"], 1, 0, "C", fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

    # --- Footer ---
    pdf.ln(10)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(130, 130, 130)
    pdf.cell(0, 5, "This report was generated by the SupportSages Linux Training Portal.", ln=True)
    pdf.cell(0, 5, "For ISO compliance and HR records.", ln=True)

    buf = BytesIO(pdf.output())
    filename = f"Training_Report_{t.full_name or t.username}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Excel Report ────────────────────────────────────────────────────────────

@router.get("/report/{trainee_id}/excel")
def report_excel(
    trainee_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    d = _trainee_report_data(trainee_id, db)
    t = d["trainee"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Training Report"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=16)
    sub_font = Font(name="Calibri", size=11, color="666666")
    bold = Font(name="Calibri", bold=True, size=11)
    normal = Font(name="Calibri", size=11)
    th_fill = PatternFill(start_color="282828", end_color="282828", fill_type="solid")
    th_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    pass_font = Font(name="Calibri", bold=True, size=10, color="008000")
    fail_font = Font(name="Calibri", bold=True, size=10, color="CC0000")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Training Completion Report - SupportSages"
    ws["A1"].font = header_font

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}"
    ws["A2"].font = sub_font

    # Trainee info
    row = 4
    ws.cell(row=row, column=1, value="Trainee Name:").font = bold
    ws.cell(row=row, column=2, value=t.full_name or t.username).font = normal
    row += 1
    ws.cell(row=row, column=1, value="Username:").font = bold
    ws.cell(row=row, column=2, value=t.username).font = normal
    if t.server_ip:
        row += 1
        ws.cell(row=row, column=1, value="Training Server:").font = bold
        ws.cell(row=row, column=2, value=t.server_ip).font = normal

    # Overall score
    row += 2
    ws.cell(row=row, column=1, value="Overall Score:").font = bold
    ws.cell(row=row, column=2, value=f"{d['total_earned']} / {d['total_max']}  ({d['total_percent']}%)").font = normal
    row += 1
    ws.cell(row=row, column=1, value="Questions Correct:").font = bold
    ws.cell(row=row, column=2, value=f"{d['total_correct']} / {d['total_questions']}").font = normal
    row += 1
    ws.cell(row=row, column=1, value="Result:").font = bold
    overall_status = "PASS" if d["total_percent"] >= 60 else "FAIL"
    c = ws.cell(row=row, column=2, value=overall_status)
    c.font = pass_font if overall_status == "PASS" else fail_font

    # Table header
    row += 2
    table_headers = ["Day", "Module", "Questions", "Correct", "Score", "%", "Status"]
    for col, h in enumerate(table_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = th_font
        c.fill = th_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # Table data
    for r in d["rows"]:
        row += 1
        vals = [r["day"], r["title"], r["questions"], r["correct"],
                f"{r['earned']}/{r['max']}", f"{r['percent']}%", r["status"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = normal
            c.border = thin_border
            c.alignment = Alignment(horizontal="center") if col != 2 else Alignment(horizontal="left")
        # Color the status column
        status_cell = ws.cell(row=row, column=7)
        status_cell.font = pass_font if r["status"] == "Pass" else fail_font

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Training_Report_{t.full_name or t.username}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Export ALL Trainees (single Excel) ──────────────────────────────────────

@router.get("/export-all-trainees")
def report_all_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    trainees = db.query(models.User).filter(models.User.role == "trainee").all()
    modules = db.query(models.Module).order_by(models.Module.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "All Trainees"

    bold = Font(name="Calibri", bold=True, size=11)
    normal = Font(name="Calibri", size=11)
    th_fill = PatternFill(start_color="282828", end_color="282828", fill_type="solid")
    th_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    pass_font = Font(name="Calibri", bold=True, size=10, color="008000")
    fail_font = Font(name="Calibri", bold=True, size=10, color="CC0000")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = "SupportSages - All Trainees Training Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16)
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")

    # Headers: Name, then each module name, then Total, %
    row = 4
    headers = ["Trainee"] + [m.title for m in modules] + ["Total Score", "Max Score", "%", "Result"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = th_font
        c.fill = th_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin_border

    for trainee in trainees:
        row += 1
        ws.cell(row=row, column=1, value=trainee.full_name or trainee.username).font = normal
        ws.cell(row=row, column=1).border = thin_border

        total_earned = 0.0
        total_max = 0.0
        for mi, m in enumerate(modules):
            m_max = sum(q.points for q in m.questions)
            m_earned = 0.0
            for q in m.questions:
                att = db.query(models.Attempt).filter(
                    models.Attempt.user_id == trainee.id,
                    models.Attempt.question_id == q.id
                ).order_by(models.Attempt.attempted_at.desc()).first()
                if att and att.is_correct:
                    m_earned += att.score or 0
            pct = round((m_earned / m_max * 100) if m_max else 0, 1)
            c = ws.cell(row=row, column=mi + 2, value=f"{pct}%")
            c.font = pass_font if pct >= 60 else fail_font
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border
            total_earned += m_earned
            total_max += m_max

        col_offset = len(modules) + 2
        ws.cell(row=row, column=col_offset, value=round(total_earned, 1)).font = normal
        ws.cell(row=row, column=col_offset).border = thin_border
        ws.cell(row=row, column=col_offset + 1, value=round(total_max, 1)).font = normal
        ws.cell(row=row, column=col_offset + 1).border = thin_border
        total_pct = round((total_earned / total_max * 100) if total_max else 0, 1)
        ws.cell(row=row, column=col_offset + 2, value=f"{total_pct}%").font = normal
        ws.cell(row=row, column=col_offset + 2).border = thin_border
        result = "PASS" if total_pct >= 60 else "FAIL"
        c = ws.cell(row=row, column=col_offset + 3, value=result)
        c.font = pass_font if result == "PASS" else fail_font
        c.border = thin_border

    # Auto-width
    ws.column_dimensions["A"].width = 25
    for i in range(len(modules)):
        ws.column_dimensions[chr(66 + i) if i < 25 else "AA"].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="All_Trainees_Report_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'},
    )


# ── Certificate PDF ─────────────────────────────────────────────────────────

@router.get("/certificate/{trainee_id}")
def generate_certificate(
    trainee_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    from fpdf import FPDF

    def _clean(text):
        return (text or "").replace("\u2014", "-").replace("\u2013", "-").replace("\u2018", "'").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"')

    d = _trainee_report_data(trainee_id, db)
    t = d["trainee"]

    if d["total_percent"] < 60:
        raise HTTPException(status_code=400, detail="Trainee has not passed (requires >= 60%)")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    # Border
    pdf.set_draw_color(0, 102, 204)
    pdf.set_line_width(2)
    pdf.rect(10, 10, 277, 190)
    pdf.set_line_width(0.5)
    pdf.rect(14, 14, 269, 182)

    # Logo
    logo_png = Path(__file__).resolve().parent.parent / "logo.png"
    if logo_png.exists():
        pdf.image(str(logo_png), x=118, y=16, w=60)
        pdf.set_xy(0, 38)
    else:
        pdf.set_xy(0, 30)

    # Header
    pdf.set_font("Helvetica", "B", 32)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(297, 15, "CERTIFICATE OF COMPLETION", 0, 1, "C")

    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(297, 8, "SupportSages - Linux Training Portal", 0, 1, "C")

    # Divider
    pdf.set_draw_color(0, 102, 204)
    pdf.set_line_width(0.5)
    pdf.line(80, 60, 217, 60)

    # Body
    pdf.set_xy(0, 70)
    pdf.set_font("Helvetica", "", 14)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(297, 10, "This is to certify that", 0, 1, "C")

    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(297, 18, _clean(t.full_name or t.username), 0, 1, "C")

    pdf.set_font("Helvetica", "", 14)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(297, 10, "has successfully completed the", 0, 1, "C")

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(297, 12, "Basic Linux System Administration Training", 0, 1, "C")

    pdf.set_font("Helvetica", "", 13)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(297, 10, f"with an overall score of {d['total_percent']}%", 0, 1, "C")
    pdf.cell(297, 8, f"({d['total_correct']}/{d['total_questions']} questions correct across 11 training modules)", 0, 1, "C")

    # Date
    pdf.set_xy(0, 155)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(297, 8, f"Date: {datetime.utcnow().strftime('%d %B %Y')}", 0, 1, "C")

    # Signature line
    pdf.set_xy(80, 170)
    pdf.set_draw_color(150, 150, 150)
    pdf.line(80, 178, 140, 178)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.set_xy(80, 179)
    pdf.cell(60, 5, "Trainer Signature", 0, 0, "C")

    pdf.set_xy(157, 170)
    pdf.line(157, 178, 217, 178)
    pdf.set_xy(157, 179)
    pdf.cell(60, 5, "Date", 0, 0, "C")

    buf = BytesIO(pdf.output())
    filename = f"Certificate_{_clean(t.full_name or t.username).replace(' ', '_')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Leaderboard ─────────────────────────────────────────────────────────────

@router.get("/leaderboard")
def leaderboard(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    trainees = db.query(models.User).filter(models.User.role == "trainee").all()
    all_questions = db.query(models.Question).all()
    max_pts = sum(q.points for q in all_questions)

    result = []
    for trainee in trainees:
        earned = 0.0
        correct = 0
        attempted = 0
        for q in all_questions:
            att = (
                db.query(models.Attempt)
                .filter(models.Attempt.user_id == trainee.id, models.Attempt.question_id == q.id)
                .order_by(models.Attempt.attempted_at.desc())
                .first()
            )
            if att:
                attempted += 1
                if att.is_correct:
                    correct += 1
                    earned += att.score or 0

        result.append({
            "trainee_id": trainee.id,
            "name": trainee.full_name or trainee.username,
            "score": round(earned, 1),
            "max_score": round(max_pts, 1),
            "percent": round((earned / max_pts * 100) if max_pts else 0, 1),
            "correct": correct,
            "attempted": attempted,
            "total_questions": len(all_questions),
        })

    # Sort by percent descending
    result.sort(key=lambda x: x["percent"], reverse=True)
    return result
